import json
from time import sleep

import openpyxl
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

wb=openpyxl.load_workbook('data.xlsx')
sheet=wb.active
cell_data=sheet.cell(row=1,column=1)
print(cell_data.value)
max_col=sheet.max_column

for i in range(1, max_col+1):
    data=sheet.cell(row=1,column=2)
    print(data.value)
