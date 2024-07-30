import json
from time import sleep

import openpyxl
import pytest
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


def verify_user(username,pass_key):
    options = Options()
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    driver.get('https://www.saucedemo.com')
    driver.maximize_window()
    sleep(5)
    user_name = driver.find_element(By.ID, 'user-name')
    password = driver.find_element(By.ID, 'password')
    login_button = driver.find_element(By.ID, 'login-button')

    user_name.send_keys(username)
    password.send_keys(pass_key)

    login_button.send_keys(Keys.ENTER)
    if driver.current_url == 'https://www.saucedemo.com/inventory.html':
        pass
    else:
        filename = "Assignment1_ScreenShots/"+username+'.png'
        driver.get_screenshot_as_file(filename)
        sleep(5)

    driver.quit()

    workBook = openpyxl.load_workbook('Assignment_data.xlsx')
    sheet = workBook.active
    max_col = sheet.max_column
    max_row = sheet.max_row

    for i in range(1, max_row + 1):
        user_name_value = sheet.cell(row=i, column=1)
        password_Value = sheet.cell(row=i, column=2)
        # print(user_name_value.value)
        # print(password_Value.value)
        verify_user(user_name_value.value, password_Value.value)




